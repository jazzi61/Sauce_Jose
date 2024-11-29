import os
from openpyxl import load_workbook
import pandas as pd
import pyodbc
from datetime import datetime

# Step 1: Define folder paths
#source_folder = 'C:\\Users\\Administrator\\Downloads\\List\\X\\Active'

#current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
#processed_folder = f'C:\\Users\\Administrator\\Downloads\\Processed_{current_datetime}'



# Ensure the processed folder exists
#if not os.path.exists(processed_folder):
#    os.makedirs(processed_folder)

# Step 2: Define the connection to SQL Server (reuse the connection)
conn = pyodbc.connect(
  'DRIVER=ODBC Driver 17 for SQL Server;'
    'SERVER=LAPTOP-3VP9184U\\SAUCESERVER;'
    'DATABASE=Raw_Data;'
    'UID=sa;'
    'Pwd=sauce123;'
    'Trusted_Connection=no;'
)
cursor = conn.cursor()


def execute_sql_job(job_name):
    try:
        sql = f"EXEC msdb.dbo.sp_start_job @job_name = '{job_name}'"
        cursor.execute(sql)
        print(f"Job '{job_name}' has been started successfully.")
    except pyodbc.Error as e:
        print(f"Error starting job '{job_name}': {str(e)}")

# Start the SQL Server job before proceeding
job_name = 'cleanjob'
#execute_sql_job(job_name)



def get_source_folders():
    query = "SELECT SourceFolderPath, Type FROM SourceFolders where IsActive = 1  and type = 'Sauce1'"
    cursor.execute(query)
    return cursor.fetchall()


# Function to map pandas data types to SQL Server data types
def map_dtype_to_sql(dtype):
    if pd.api.types.is_integer_dtype(dtype):
        return 'INT'
    elif pd.api.types.is_float_dtype(dtype):
        return 'FLOAT'
    elif pd.api.types.is_bool_dtype(dtype):
        return 'BIT'
    elif pd.api.types.is_datetime64_any_dtype(dtype):
        return 'DATETIME'
    else:
        return 'VARCHAR(700)'



# Function to get columns from a template table
def get_template_columns(template_table_name):
    query = f"SELECT Columns FROM dbo.{template_table_name} WHERE status = 1"
    #print(query)
    cursor.execute(query)
    return {row[0] for row in cursor.fetchall()}

# Function to add new columns to the SQL table (if any new columns are required)
def add_new_columns_to_table(table_name, new_columns):
    for col in new_columns:
        sql_type = map_dtype_to_sql(pd.Series(dtype='object'))  # Default to VARCHAR for new columns
        alter_table_query = f"ALTER TABLE {table_name} ADD [{col}] {sql_type};"
        try:
            cursor.execute(alter_table_query)
            print(f"Added column '{col}' to table '{table_name}'.")
        except Exception as e:
            print(f"Error adding column '{col}' to table '{table_name}': {str(e)}")



# Special keywords for table name logic
special_keywords = ['Probate', 'Tax', 'Eviction']

def log_insertion_error(file_name, sheet_name, row_index, error_message):
    error_insert_query = '''
        INSERT INTO dbo.Insertion_Errors (File_Name, Sheet_Name, Row_Index, Error_Message)
        VALUES (?, ?, ?, ?)
    '''
    cursor.execute(error_insert_query, file_name, sheet_name, row_index, error_message)
    conn.commit()

# Function to check if a table exists
def table_exists(table_name):
    check_table_query = f"SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{table_name}'"
    cursor.execute(check_table_query)
    return cursor.fetchone()[0] > 0

# Function to get current columns in the SQL table
def get_current_columns(table_name):
    query = f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{table_name}'"
    cursor.execute(query)
    return {row[0] for row in cursor.fetchall()}

# Function to add new columns to the SQL table
def add_new_columns_to_table(table_name, new_columns):
    for col in new_columns:
        sql_type = map_dtype_to_sql(pd.Series(dtype='object'))  # Default to VARCHAR for new columns
        alter_table_query = f"ALTER TABLE {table_name} ADD [{col}] {sql_type};"
        try:
            cursor.execute(alter_table_query)
            #print(f"Added column '{col}' to table '{table_name}'.")
        except Exception as e:
            print(f"Error adding column '{col}' to table '{table_name}': {str(e)}")

# Batch insert function to reduce DB operations
def batch_insert_to_sql(df, insert_query, table_name):
    batch_size = 2000  # Number of rows to insert at a time
    total_rows = len(df)
    for start in range(0, total_rows, batch_size):
        batch_df = df[start:start + batch_size]
        try:
            cursor.fast_executemany = True
            cursor.executemany(insert_query, batch_df.values.tolist())
        except Exception as e:
            log_insertion_error(table_name, file_name, start, str(e))
    conn.commit()


# Special keywords mapped to corresponding template tables
keyword_to_template_table = {
    'Sauce3': 'Sauce3Template',
    'Sauce1': 'Sauce1Template',
    'Sauce2': 'Sauce2Template',
    # Add other mappings here if necessary
}

# Function to get the template table based on the file name
def get_template_table_name(file_type):
    for keyword, template_table in keyword_to_template_table.items():
        if keyword.lower() in file_type.lower():
            return template_table
    return None  # Return None if no keyword matches

#def is_file_processed(file_path):
#    normalized_file_path = os.path.normpath(file_path)  # Normalize for comparison
#    return normalized_file_path in processed_files_log

# Function to log processed file in both the database and in-memory set
def log_processed_file(file_name, file_path, IsProcessed):
    cursor.execute("INSERT INTO ProcessedFilesLog (File_Name, File_Path, IsProcessed) VALUES (?, ?, ?)", (file_name, file_path, IsProcessed))
    conn.commit()
    processed_files_log.add(os.path.normpath(file_path))  # Also log in memory


# Function to manually chunk large DataFrames
def chunk_df(df, chunk_size=1000):
    for start in range(0, df.shape[0], chunk_size):
        yield df.iloc[start:start + chunk_size]

def get_table_name(file_name, sheet_name):
    # Look for special keywords in the file name
    for keyword in special_keywords:
        if keyword.lower() in file_name.lower():
            table_name = f'{keyword}'.replace(" ", "_")
            #print(f"Special keyword '{keyword}' found in file name. Table name set to: {table_name}")
            return table_name
    # Default table name if no keyword found
    #table_name = f'{file_name.replace(".xlsx", "")}_{sheet_name}'.replace(" ", "_")
    #print(f"No special keyword found. Default table name: {table_name}")
    return table_name

table_name = 'null'

def makeProcessedFolder(foldername):
    current_datetime = datetime.now().strftime('%Y-%m-%d_%H-')
    processed_folder = f'C:\\log\\Processed_{current_datetime}_{foldername}'
    #print('Making Folder' + processed_folder )
    if not os.path.exists(processed_folder):
        os.makedirs(processed_folder,exist_ok=False)
    return processed_folder

# Existing extract_clickable_text_from_hyperlink function (unchanged)
def extract_clickable_text_from_hyperlink(sheet):
    # Create a list to store row data
    data = []
    hyperlinks = []  # Store hyperlinks if found

    # Iterate through each row in the sheet, skipping the header row (index 1)
    for row in sheet.iter_rows(min_row=2, values_only=False):  # Start from the second row
        row_data = []
        hyperlink_data = []  # Store hyperlink details row-wise

        # Check if the row is empty and skip it
        if all(cell.value is None for cell in row):  # If all cell values are None
            continue  # Skip this empty row
        
        for cell in row:
            cell_value = cell.value if cell.value is not None else ''  # Get cell value, default to empty string
            hyperlink = None  # Default to None if no hyperlink

            # Check if the cell contains a HYPERLINK formula
            if isinstance(cell_value, str) and cell_value.startswith('=HYPERLINK'):
                # Extract the display text from the formula
                try:
                    # Split the string to isolate the display text
                    parts = cell_value.split('",')
                    if len(parts) == 2:
                        display_text = parts[1].strip().strip('")"')  # Get the second part (display text)
                        hyperlink = parts[0].strip('=HYPERLINK(")').strip('"')  # Store the actual hyperlink
                    elif len(parts) > 2:
                        display_text = cell_value  # Get the second part (display text)
                        hyperlink = parts[0].strip('=HYPERLINK(")').strip('"')  # Store the actual hyperlink
                    else:
                        display_text = cell_value  # Fallback to the whole value if the format is unexpected
                except Exception as e:
                    display_text = cell_value  # Fallback to the whole value on error
            else:
                display_text = cell_value  # Use cell value as is

            # Append the display text (or cell value) to the row data
            row_data.append(display_text)

            # Append the hyperlink to hyperlink_data if extracted
            hyperlink_data.append(hyperlink)

        # Add the row data and hyperlink data to respective lists if they're not empty
        if row_data:  # Ensure we're only adding non-empty rows
            data.append(row_data)
            hyperlinks.append(hyperlink_data)
            
    return data, hyperlinks


processed_files_log = set()

def setProcessedFile():
    processed_files_log = set()
# Before processing, load already processed files from a log table (or file)
    cursor.execute("SELECT File_Path FROM ProcessedFilesLog")
    for row in cursor.fetchall():
        processed_files_log.add(row[0])

def is_file_processed(file_path):
    try:
        # Normalize the file path for comparison
        normalized_file_path = os.path.normpath(file_path)

        # Query to check if the file exists in the ProcessedFilesLog table
        query = "SELECT COUNT(*) FROM ProcessedFilesLog WHERE File_Path = ?"
        cursor.execute(query, normalized_file_path)
        result = cursor.fetchone()

        # If the result is greater than 0, the file has been processed
        return result[0] > 0
    except pyodbc.Error as e:
        print(f"Error checking if file '{file_path}' has been processed: {str(e)}")
        return False

# Process each file based on folder and type from SQL
for source_folder, file_type in get_source_folders():
    print(f"Outer Loop in folder: {source_folder} of type: {file_type}")

    for dirpath, _, filenames in os.walk(source_folder):
        for file_name in filenames:
            if file_name.endswith('.xlsx') : 

                file_path = os.path.join(dirpath, file_name)
                # Check if the file has already been processed (using in-memory log)
                if is_file_processed(file_path):
                    print(f"File '{file_path}' has already been processed in this session. Skipping.")
                    continue  # Skip this file

                
                # Step 5: Determine template table based on the file name
                template_table_name = get_template_table_name(file_type)
                if not template_table_name:
                    print(f"No matching keyword found in the file name '{file_name}'. Skipping file.")
                    continue

                # Mpved Down in Code Step 6: Get the columns from the identified template table
                # template_columns = get_template_columns(template_table_name)
                
                # Step 7: Open and process the Excel file
                try:
                    sheets_dict = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                    for sheet_name, df in sheets_dict.items():
                        #print(f"Processing sheet: {sheet_name}")

                        # Generate table name based on file name and sheet name
                        table_name = file_type #get_table_name(file_name, sheet_name)
                        
                        wb = load_workbook(file_path)
                        sheet = wb[sheet_name]  # Specify the sheet you want to work with
                        # Step 3: Extract clickable text from the sheet's hyperlinks
                        data, hyperlinks = extract_clickable_text_from_hyperlink(sheet)

                        # Step 2: Convert the extracted data into a pandas DataFrame
                        newdf = pd.DataFrame(data, columns=df.columns)

                        # Step 3: Create a DataFrame for hyperlinks (if any were extracted)
                        hyperlinks_df = pd.DataFrame(hyperlinks, columns=[col + "_hyperlink" for col in df.columns])

                        # Step 4: Concatenate the new DataFrame with the hyperlinks DataFrame
                        combined_df = pd.concat([newdf, hyperlinks_df], axis=1)
                        df = combined_df

                        df.columns = [col.replace(' - ', ' ').replace(' ', '_').replace('$', '')
                                    .replace('-', '_').replace('.', '').replace('#', 'No')
                                    .replace('Use', 'Uses').replace('(', '').replace(')', '').replace('__', '_')
                                    for col in df.columns]
                        #df = df.replace(np.nan, None)
                        #df = df.replace({float('nan'): None})  # Replace NaN values
                        #df = df.fillna('')
                        #print(df.columns)
                        # Skip empty sheets

                        if df.empty:
                            #print(f"Sheet '{sheet_name}' is empty. Skipping.")
                            continue

                        # Capture current datetime for bulk insert
                        bulk_insert_datetime = datetime.now()

                        # Prepare the new columns as a dictionary
                        new_columns = {
                            'File_Name': file_name,
                            'Sheet_Name': sheet_name,
                            'Bulk_Insert_DateTime': bulk_insert_datetime,
                            'IsShifted': 0
                        }

                        # Create a DataFrame from the new columns
                        new_columns_df = pd.DataFrame(new_columns, index=df.index)

                        # Concatenate the new columns DataFrame with the original DataFrame
                        df = pd.concat([df, new_columns_df], axis=1)

                        # Optional: Create a copy to de-fragment the DataFrame
                        df = df.copy()

                        
                        # Step 6: Get the columns from the identified template table
                        template_columns = get_template_columns(template_table_name)
                        
                        # Filter DataFrame to keep only columns that exist in the template table
                        filtered_columns = [col for col in df.columns if col in template_columns]
                        df = df[filtered_columns]

                        # Skip if no columns match the template table
                        if df.empty:
                            print(f"No matching columns in sheet '{sheet_name}'. Skipping.")
                            continue
                        
                        
                        # Step 6: Check if the table already exists
                        if not table_exists(table_name):
                            print(f"Creating new table: {table_name}")

                            # Dynamically generate SQL table creation statement
                            sql_columns = []
                            for col in df.columns:
                                sql_type = map_dtype_to_sql(df[col].dtype)
                                sql_columns.append(f'[{col}] {sql_type}')

                            create_table_statement = f'''
                            CREATE TABLE {table_name} (
                                {', '.join(sql_columns)}
                            );
                            '''
                            cursor.execute(create_table_statement)
                        else:
                            # Get current columns in the SQL table
                            existing_columns = get_current_columns(table_name)
                            # Identify new columns in the DataFrame
                            new_columns = set(df.columns) - existing_columns
                            # Add new columns to the SQL table if they do not exist
                            if new_columns:
                                #print(f"NEW COLUMNS FOUND'{sheet_name}'")
                                add_new_columns_to_table(table_name, new_columns)

                        # Prepare SQL insert query

                        placeholders = ', '.join('?' * len(df.columns))
                        columns_str = ', '.join([f'[{col}]' for col in df.columns])

                        

                        insert_query = f'''
                            INSERT INTO {table_name} ({columns_str})
                            VALUES ({placeholders})
                        '''
                        #print (insert_query)
               
                        # Step 7: Insert data in chunks
                        for chunk in chunk_df(df):
                            batch_insert_to_sql(chunk, insert_query, table_name)


                    log_processed_file(file_name, file_path,1)
                    print(f"File '{file_path}'  Loaded.")
                    #print("Moving File")
                    # Step 8: Move the processed file to the "Processed" folder
                    #processed_folder = makeProcessedFolder(table_name)
                    #processed_subfolder = os.path.join(processed_folder, os.path.relpath(dirpath, source_folder))
                    #if not os.path.exists(processed_subfolder):
                    #    os.makedirs(processed_subfolder)

                    #shutil.move(file_path, os.path.join(processed_subfolder, file_name))
                    #print(f"File '{file_name}' processed and moved to '{processed_subfolder}'.")

                    
                
                except Exception as e:
                    print(f"File '{file_path}'  Reload Later.")
                    log_insertion_error(table_name, file_name, '0', str(e))
                    #log_processed_file(file_name, file_path,0)
                    continue
        
            

        

# Step 9: Close the database connection
cursor.close()
conn.close()
